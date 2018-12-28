import openpyxl

wb1 = openpyxl.load_workbook('apps.xlsx')
#wb2 = openpyxl.load_workbook('awsdm.xlsx')

ws1 = wb1.active
#ws2 = wb2.active

alert_announce_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/alert-announce/"
alert_announce_module = "alert_announce"
alert_announce_worksheet = wb1.create_sheet(title = alert_announce_module)

awsdm_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/awsdm/1.0-r0/fulcrum/awsdm/"
awsdm_module = "awsdm"
awsdm_worksheet = wb1.create_sheet(title = awsdm_module)

bbrpc_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/bbrpc/"
bbrpc_module = "bbrpc"
bbrpc_worksheet = wb1.create_sheet(title = bbrpc_module)

fulcrum_voip_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/fulcrum-voip/"
fulcrum_voip_module = "fulcrum_voip"
fulcrum_voip_worksheet = wb1.create_sheet(title = fulcrum_voip_module)

gui_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/oem-gui/"
gui_module = "gui"
gui_worksheet = wb1.create_sheet(title = gui_module)

fota_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/fota/"
fota_module = "fota"
fota_worksheet = wb1.create_sheet(title = fota_module)

get_hwid_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/get-hwid/"
get_hwid_module = "get_hwid"
get_hwid_worksheet = wb1.create_sheet(title = get_hwid_module)

mediaserver_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/oem-mediaserver/"
mediaserver_module = "mediaserver"
mediaserver_worksheet = wb1.create_sheet(title = mediaserver_module)

sscep_client_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/sscep-client/"
sscep_client_module = "sscep_client"
sscep_client_worksheet = wb1.create_sheet(title = sscep_client_module)

diag_log_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/diag-log/"
diag_log_module = "diag_log"
diag_log_worksheet = wb1.create_sheet(title = diag_log_module)

diagtool_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/diagtool/"
diagtool_module = "diagtool"
diagtool_worksheet = wb1.create_sheet(title = diagtool_module)

diagnostic_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/diagnostic/"
diagnostic_module = "diagnostic"
diagnostic_worksheet = wb1.create_sheet(title = diagnostic_module)

error_handle_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/err-handle/"
error_handle_module = "error_handle"
error_handle_worksheet = wb1.create_sheet(title = error_handle_module)

batpersent_path = "poky/build/tmp-glibc/work/armv7a-vfp-neon-oe-linux-gnueabi/batpersent/"
batpersent_module = "batpersent"
batpersent_worksheet = wb1.create_sheet(title = batpersent_module)

def iter_rows(ws,n):  #produce the list of items in the particular row
        for row in ws.iter_rows(n):
            yield [cell.value for cell in row]
            
def create_sheet_for_mot_klockwork(worksheet,sheet_module,cell):
    n= 'A' + format(cell.row) + ':' + ('GH' + format(cell.row))
    list_to_append = list(iter_rows(worksheet,n))
    for items in list_to_append:
        sheet_module.append(items)
        
for row in ws1.iter_rows():
    for cell in row:
        if  alert_announce_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,alert_announce_worksheet,cell)
        elif  awsdm_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,awsdm_worksheet,cell)
        elif bbrpc_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,bbrpc_worksheet,cell)
        elif fulcrum_voip_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,fulcrum_voip_worksheet,cell)
        elif gui_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,gui_worksheet,cell)
        elif fota_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,fota_worksheet,cell)
        elif get_hwid_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,get_hwid_worksheet,cell)
        elif mediaserver_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,mediaserver_worksheet,cell)
        elif sscep_client_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,sscep_client_worksheet,cell)
        elif diag_log_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,diag_log_worksheet,cell)
        elif diagtool_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,diagtool_worksheet,cell)
        elif diagnostic_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,diagnostic_worksheet,cell)
        elif error_handle_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,error_handle_worksheet,cell)
        elif batpersent_path in cell.value:
            create_sheet_for_mot_klockwork(ws1,batpersent_worksheet,cell)

            
wb1.save("apps.xlsx")
print ("Klockwork parser is finished")