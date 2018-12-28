'''
    워크북 생성
    데이터 포맷팅
    함수사용
    셀병합
    이미지 삽입
'''

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image  # PIL 모듈 필요( pip install PIL 에러시 pip3 install pillow로 설치 )
import datetime


def write_workbook():
    print('===write_workbook===')
    wb = Workbook()
    dest_filename = 'write_workbook.xlsx'
    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
         ws1.append((1,2,3)) # 시트에 append를 하면 row가 추가된다. 이때 list, tuple에서 각 아이템은 컬럼값으로 들어간다.

    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")

    for row in range(10, 20):
         for col in range(27, 54):
             _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

    print(ws3['AA10'].value)

    wb.save(filename=dest_filename)


def number_formats():
    print('===number_formats===')
    wb = Workbook()
    ws = wb.active

    ws['A1'] = datetime.datetime(2010, 7, 21)
    print(ws['A1'].number_format)
    print(ws['A1'].value)

    wb.guess_types = True

    ws['B1'] = '3.14%'
    wb.guess_types = False
    print(ws['B1'].value)
    print(ws['B1'].number_format)

    dest_filename = 'number_formats.xlsx'
    wb.save(filename=dest_filename)


def formulae():
    print('===formulae===')

    wb = Workbook()
    ws = wb.active
    # add a simple formula

    ws["A1"] = "=SUM(1, 1)"
    wb.save("formula.xlsx")


def merge_cell():
    print('===merge_cell===')

    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A2:D2')
    ws.unmerge_cells('A2:D2')

    # or equivalently
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    # ws.unmerge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    wb.save("merge_cell.xlsx")


def insert_img():
    print('===insert_img===')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'You should see three logos below'

    img = Image('frog.png')
    ws.add_image(img, 'A2')
    wb.save("insert_img.xlsx")


def fold_column():
    print('===fold_column===')
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions.group('A', 'D', hidden=True)
    wb.save('fold_column.xlsx')

if __name__ == '__main__':
    write_workbook()
    number_formats()
    formulae()
    merge_cell()
    insert_img()
    fold_column()