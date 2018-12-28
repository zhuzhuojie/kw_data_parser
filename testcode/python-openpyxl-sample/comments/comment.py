from openpyxl import Workbook
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
comment = Comment("Text", "Author")
ws["A1"].comment = comment
ws["B2"].comment = comment
print(ws["A1"].comment is comment)

print(ws["B2"].comment is comment)

wb.save('comment.xlsx')