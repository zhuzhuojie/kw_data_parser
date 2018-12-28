"""
    폰트 스타일
    테두리 스타일
    셀 스타일
    로우 고정(맥에서는 확인이 안됨...)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color


def font_style():
    print('====font_style=====')
    wb = Workbook()
    ws = wb.active

    ws['A1'].value = '안녕하세요 멍개입니다.'
    ca1 = ws['A1']

    # 폰트 이름은 '맑은 고딕'이고 크기는 15이면서 굵게 속성 설정
    ca1.font = Font(name='맑은 고딕', size=12, bold=True)
    ca1.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    wb.save('font_style.xlsx')


def border_style():
    print('====border_style=====')

    wb = Workbook()
    ws = wb.active

    ws['A1'] = '테스트'
    ca2 = ws['A1']

    # border_style : {'mediumDashed', 'dashDotDot', 'mediumDashDot', 'medium', 'thick', 'mediumDashDotDot', 'thin', 'double', 'dashDot', 'slantDashDot', 'dashed', 'dotted', 'hair'}

    box = Border(left=Side(border_style="thin",
                           color='FF000000'),
                 right=Side(border_style="thin",
                            color='FF000000'),
                 top=Side(border_style="thin",
                          color='FF000000'),
                 bottom=Side(border_style="thin",
                             color='FF000000'),
                 diagonal=Side(border_style="thin",
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style="thin",
                              color='FF000000'),
                 vertical=Side(border_style="thin",
                               color='FF000000'),
                 horizontal=Side(border_style="thin",
                                 color='FF000000')
                 )

    ca2.border = box  # Cell 테두리를 적용한다.

    wb.save('border_style.xlsx')


def cell_color():
    print('====cell_color=====')

    wb = Workbook()
    ws = wb.active

    c2 = ws['C2']
    c2.fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))

    wb.save('cell_color.xlsx')


def row_fixed():
    print('====row_fixed=====')
    wb = Workbook()
    ws = wb.active

    ws.freeze_panes = 'A2'  # A2을 고정시킨다

    wb.save('row_fixed.xlsx')


if __name__ == '__main__':
    font_style()
    border_style()
    cell_color()
    row_fixed()