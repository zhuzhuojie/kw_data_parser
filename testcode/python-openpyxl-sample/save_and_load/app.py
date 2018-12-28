'''
엑셀 저장
    엑셀에서 워크북을 생성 후 시트추가.
    시트접근 방법 및 시트이름 변경
    각 셀 접근 및 데이터 수정 방법
엑셀 읽기
    각 시트 가져오기
    데이터 접근
'''

from openpyxl import Workbook, load_workbook


def save_document(filename):
    wb = Workbook()      # 워크북을 생성한다.
    ws = wb.active       # 워크 시트를 얻는다.
    ws['B2'] = "asdflkjasdlkfj"

    ws1 = wb.create_sheet('sheet1', 0)       # 워크 시트를 얻는다.
    ws1['A1'] = 'Hello1'   # A1에 'Hello' 값을 입력한다.

    ws2 = wb.create_sheet('sheet2', 1)       # 워크 시트를 얻는다.
    ws2['A1'] = 'Hello2'   # A1에 'Hello' 값을 입력한다.

    print(wb.sheetnames)  # sheet들의 이름을 리스트 형태로 가져온다.

    for sheet in wb:      # 워크북에서 시트들을 가져온다.
        print(sheet.title)

    ws3 = wb["sheet1"]  # sheet copy
    ws3['B5'] = '안녕'   # sheet1는 ws1과 ws3으로 조작가능

    d = ws.cell(row=4, column=2, value=10)  # row와 column으로 값 수정 : 이때는 0이 아닌 1이 시작값
    print(d)   # 셀 객체에서 value 속성으로 해당셀의 값을 가져올 수 있다.

    cell_range = ws['A1':'C2']  # 셀도 슬라이싱이 가능
    print(cell_range)

    colC = ws['C']        # row 또는 column 단위로 인덱싱 및 슬라이싱 가능 -> pandas와 비슷
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]

    print(colC)
    print(col_range)
    print(row10)
    print(row_range)

    for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):  # 컬럼몇과 로우명이 아닌 숫자로 루프가능 : 기준 row
        for cell in row:
            print(cell)

    for row in ws.iter_cols(min_row=1, max_col=3, max_row=2):  # 컬럼몇과 로우명이 아닌 숫자로 루프가능 : 기준 column
        for cell in row:
            print(cell)

    ws1.title = "mmmm" # sheet 이름 변경

    wb.save(filename)  # 엑셀로 저장한다.


def load_document(filename):
    wb = load_workbook(filename)
    sheets = {}

    for sheet in wb.get_sheet_names():
        print(sheet)
        sheets[sheet] = wb[sheet]

    print(sheets['sheet2']['A1:G5'])  # 각 시트를 가져온 후 슬라이싱으로 데이터 접근해보기


if __name__ == "__main__":
    filename = 'test.xlsx'

    save_document(filename)
    load_document(filename)