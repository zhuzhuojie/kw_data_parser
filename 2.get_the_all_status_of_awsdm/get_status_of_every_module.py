# -*- coding: utf-8 -*-
"""
Created on Fri Dec 28 14:48:48 2018

@author: zhuzhuojie
"""

import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo

wb1 = openpyxl.load_workbook('apps_awsdm.xlsx')

ws1 = wb1.active

critical_number = 0
error_number = 0
warning_number = 0
review_number = 0

for row in ws1.iter_rows():
    for cell in row :
        if cell.column == 'B':
            if 'Critical' in cell.value:
                critical_number = critical_number + 1
                print ("### col :"+cell.column)
            elif  'Error' in cell.value:
                error_number = error_number + 1
            elif  'Warning' in cell.value:
                warning_number = warning_number + 1    
            elif  'Review' in cell.value:
                review_number = review_number + 1    

print ("## awsdm critical_number :"+ format(critical_number))
print ("## awsdm error_number :"+ format(error_number))
print ("## awsdm warning_number :"+ format(warning_number))
print ("## awsdm review_number :"+ format(review_number))

summary_module = "summary"

summary_worksheet = wb1.create_sheet(title = summary_module)

# add column headings. NB. these must be strings
summary_worksheet.append(["module", "(1)Critical", "(2)Error", "(3)Warning", "(4)Review"])
summary_worksheet.append(["awsdm", critical_number, error_number, warning_number, review_number])

tab = Table(displayName="Table1", ref="A1:E2")

# Add a default style with striped rows and banded columns
#style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                       showLastColumn=True, showRowStripes=False, showColumnStripes=True)

tab.tableStyleInfo = style
summary_worksheet.add_table(tab)

wb1.save("apps_awsdm.xlsx")

print ("apps awsdm is finished")