# -*- coding: utf-8 -*-
"""
Created on Fri Dec 28 16:12:33 2018

@author: zhuzhuojie
"""

import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo

wb1 = openpyxl.load_workbook('apps_all_module.xlsx')

sheets = {}



ws1 = wb1.active

summary_module = "summary"

summary_worksheet = wb1.create_sheet('summary',0)

summary_worksheet.append(["module", "(1)Critical", "(2)Error", "(3)Warning", "(4)Review"])
tab = Table(displayName="Table1", ref="A1:E15")


def add_number_for_module(worksheet,sheet_module,summary):
    critical_number = 0
    error_number = 0
    warning_number = 0
    review_number = 0

    for row in worksheet.iter_rows():
        for cell in row :
            if cell.column == 'B':
                print ("## cell.value:"+cell.value)
                if 'Critical' in cell.value:
                    critical_number = critical_number + 1
                elif  'Error' in cell.value:
                    error_number = error_number + 1
                elif  'Warning' in cell.value:
                    warning_number = warning_number + 1    
                elif  'Review' in cell.value:
                    review_number = review_number + 1    

    print ("## critical_number :"+ format(critical_number))
    print ("## error_number :"+ format(error_number))
    print ("## warning_number :"+ format(warning_number))
    print ("## review_number :"+ format(review_number))

    summary.append([sheet_module, critical_number, error_number, warning_number, review_number])


for sheet in wb1.get_sheet_names():
    print("## sheet : "+sheet)
    sheets[sheet] = wb1[sheet]
    if 'summary' not in sheet:
        add_number_for_module(sheets[sheet],sheet,summary_worksheet)



# Add a default style with striped rows and banded columns
#style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
#                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                       showLastColumn=True, showRowStripes=False, showColumnStripes=True)

tab.tableStyleInfo = style
summary_worksheet.add_table(tab)


wb1.save("apps_all_module.xlsx")

print ("get all summery is finished")